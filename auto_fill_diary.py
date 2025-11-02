# auto_fill_diary.py
# Excel auto-fill processor with "zero-update guard" and minimal logs.
from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime, date, timezone, timedelta
import re
from typing import Optional, Tuple, Dict, Any

JST = timezone(timedelta(hours=9))

TARGET_COLS = ['V','X','Y','AA']
HOLIDAY_COL = 'U'
DATE_COL = 'A'

def _normalize_digits(s: str) -> str:
    return s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))

def _match_target_sheet(title: str, m: int) -> bool:
    t = _normalize_digits(title)
    return re.search(rf"日誌\s*0?{m}\s*月\s*$", t) is not None

def _extract_month_from_title(title: str) -> Optional[int]:
    t = _normalize_digits(title)
    m = re.search(r"(\d{1,2})\s*月", t)
    if m:
        v = int(m.group(1))
        if 1 <= v <= 12:
            return v
    return None

def _is_whitespace_or_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return re.match(r"^\s*$", v) is not None
    return False

def _val_equals_50_any(v: Any) -> bool:
    # Accept numeric 50, string "50" (half/full width), and strings with spaces
    if v is None:
        return False
    if isinstance(v, (int, float)):
        try:
            return float(v) == 50.0
        except Exception:
            return False
    s = str(v)
    s = _normalize_digits(s).strip()
    return s == "50"

def _coerce_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Try parse "YYYY/M/D" or "M/D" or "YYYY-MM-DD"
    s = str(v).strip()
    s = s.replace("年","/").replace("月","/").replace("日","").replace("-", "/")
    s = re.sub(r"/+", "/", s)
    for fmt in ("%Y/%m/%d","%m/%d","%Y/%-m/%-d","%Y/%m/%-d","%Y/%-m/%d"):
        try:
            dt = datetime.strptime(s, fmt)
            # If year omitted, assume this year
            if fmt == "%m/%d":
                dt = dt.replace(year=datetime.now(JST).year)
            return dt.date()
        except Exception:
            continue
    return None

def _cell_is_writable(ws, cell) -> bool:
    # If sheet protection is ON, require cell.protection.locked == False
    prot_on = bool(getattr(ws.protection, "sheet", False))
    if not prot_on:
        return True
    cp = getattr(cell, "protection", None)
    return (cp is not None) and (cp.locked is False)

def _above_known_value(ws, r: int, c: int) -> Optional[Any]:
    # Scan upwards for the nearest non-empty value in this column
    for rr in range(r-1, 0, -1):
        v = ws.cell(row=rr, column=c).value
        if not _is_whitespace_or_empty(v):
            return v
    return None

def process(path: str) -> Tuple[int, Dict[str, Any]]:
    """Open the workbook, fill empty cells in V/X/Y/AA under rules, save once.
    Returns: (modified_count, logs)"""
    wb = load_workbook(path, data_only=False, keep_vba=True)
    today = datetime.now(JST).date()
    month_num = today.month

    # Detect target sheet
    primary = None
    for ws in wb.worksheets:
        if _match_target_sheet(ws.title, month_num):
            primary = ws
            break
    if primary is None:
        # fallback: sheet containing both "日誌" and "月" with largest month
        fallback = None
        maxm = -1
        for ws in wb.worksheets:
            t = _normalize_digits(ws.title)
            if ("日誌" in t) and ("月" in t):
                mv = _extract_month_from_title(ws.title)
                if mv is not None and mv > maxm:
                    fallback, maxm = ws, mv
        target_ws = fallback or wb.active
    else:
        target_ws = primary

    # Precompute column indexes
    col_idx = {L: column_index_from_string(L) for L in TARGET_COLS + [HOLIDAY_COL, DATE_COL]}
    ws = target_ws

    total_rows_checked = 0
    holidays = 0
    empty_cells_seen = 0
    modified = 0
    evaluated_rows = 0

    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        # Date gating: only this month and <= today
        d = ws.cell(row=r, column=col_idx[DATE_COL]).value
        dco = _coerce_date(d)
        if dco is None or dco.month != month_num or dco > today:
            continue
        evaluated_rows += 1

        # Holiday gating
        uval = ws.cell(row=r, column=col_idx[HOLIDAY_COL]).value
        if _val_equals_50_any(uval):
            holidays += 1
            continue

        # For each target col: fill only empty & writable
        for L in TARGET_COLS:
            c = col_idx[L]
            cell = ws.cell(row=r, column=c)
            if not _is_whitespace_or_empty(cell.value):
                continue
            empty_cells_seen += 1

            if not _cell_is_writable(ws, cell):
                continue

            # Simple fill policy: copy nearest above known value in same column
            # If none, leave empty (respect "推測不能は空欄維持")
            fill_val = _above_known_value(ws, r, c)
            if fill_val is None:
                continue

            cell.value = fill_val
            modified += 1

    # Save once
    wb.save(path)

    logs = {
        "sheet": ws.title,
        "month": month_num,
        "today": today.isoformat(),
        "rows_evaluated": evaluated_rows,
        "holidays_skipped": holidays,
        "empty_cells_seen": empty_cells_seen,
        "modified_count": modified,
        "sheet_protection": bool(getattr(ws.protection, "sheet", False)),
    }
    return modified, logs
