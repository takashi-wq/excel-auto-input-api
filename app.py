# app.py
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Optional, List
from pydantic import BaseModel
import io, os, tempfile, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = FastAPI(title="Excel Auto Fill API", version="1.0.0")

PASSWORD_ENV = "PASSWORD"

# ---- ユーティリティ ---------------------------------------------------------
def read_password() -> Optional[str]:
    return os.getenv(PASSWORD_ENV)

def now_str() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def safe_sheet(ws_name: str) -> str:
    # レンダーログを見やすくするための短縮
    return ws_name.replace("\n", "\\n")

# ---- 実ファイル処理 ---------------------------------------------------------
def write_debug_stamps(wb) -> List[str]:
    """
    すべてのワークシートの Z1 セルにデバッグ印を押す。
    実際にどのシートに書けたか（=変更があったか）を返す。
    """
    touched: List[str] = []
    for ws in wb.worksheets:
        try:
            ws["Z1"] = f"DEBUG {now_str()}"
            touched.append(ws.title)
        except Exception as e:
            # 書けなかったシートもログに残したい
            touched.append(f"{ws.title} (write-error: {e})")
    return touched


def try_fill_template(wb) -> List[str]:
    """
    あなたの想定テンプレに対して試し書き（少量）を行う。
    どこに書いたかのログを返す。
    ※位置が違う可能性があるので、まずは無害な「右上あたり」に入れる。
    """
    logs: List[str] = []
    # 候補：名前に「日誌」「入力」「sheet」「diary」等が含まれるシートを優先
    candidates = []
    for ws in wb.worksheets:
        name = ws.title
        key = name
        rank = 0
        for token in ("日誌", "入力", "diary", "sheet", "実習"):
            if token in name:
                rank += 1
        candidates.append((rank, name))
    # ランク高い順
    candidates.sort(reverse=True)

    targets = [n for _, n in candidates[:2]] or [wb.worksheets[0].title]
    targets = list(dict.fromkeys(targets))  # unique

    for title in targets:
        try:
            ws = wb[title]
            # ※結合セルにぶつからないよう、目立たず比較的安全な XFD10 の隣近辺に書く
            ws["AA10"] = "API書込テスト：50"
            ws["AB10"] = "API書込テスト：15"
            logs.append(f"wrote AA10/AB10 in '{safe_sheet(title)}'")
        except Exception as e:
            logs.append(f"failed write in '{safe_sheet(title)}' -> {e}")
    return logs

# ---- API --------------------------------------------------------------------
@app.get("/")
def health():
    return {"status": "ok", "time": now_str()}

@app.post("/process")
async def process(
    file: UploadFile = File(..., description=".xlsx をアップロード"),
    password: Optional[str] = Form(None, description="アップロード用パスワード（環境変数と一致が必要）"),
    debug: Optional[bool] = Form(False, description="True で全シートの Z1 にデバッグ印"),
):
    # パスワードチェック
    expected = read_password()
    if expected:
        if password != expected:
            return JSONResponse(status_code=401, content={"detail": "Unauthorized"})

    # 拡張子チェック
    filename = file.filename or "uploaded.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx is supported")

    # 一時ファイルに保存（日本語名でもOK）
    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        content = await file.read()
        tmp_in.write(content)
        tmp_in.flush()
        tmp_in.close()

        # Excel を読み込んで処理
        wb = load_workbook(tmp_in.name, data_only=False)  # 書き込み可能

        # ---- ここがポイント：まず全シートにデバッグ印を打つかどうか ----
        action_logs: List[str] = []
        if debug:
            touched = write_debug_stamps(wb)
            action_logs.append(f"debug stamps -> {touched}")

        # ついでに “想定テンプレ” にも軽く書いてみる（右上あたり）
        try_logs = try_fill_template(wb)
        action_logs.extend(try_logs)

        # 保存して再読込（openpyxl の一部挙動対策）
        tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp_out.name)
        tmp_out.flush()
        tmp_out.close()

        # レスポンスへ
        output_name = os.path.splitext(os.path.basename(filename))[0] + " updated.xlsx"
        with open(tmp_out.name, "rb") as fh:
            stream = io.BytesIO(fh.read())

        # レンダーのログで何が起きたか見えるように、ヘッダにも情報を少し出す
        headers = {"X-Write-Log": "; ".join(action_logs)[:1024]}

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{output_name}"', **headers},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing failed: {e}")
    finally:
        try:
            os.remove(tmp_in.name)
        except Exception:
            pass
